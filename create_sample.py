"""샘플 정비 비용정산 Excel 파일 생성 (테스트용)"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()

bold14 = Font(bold=True, size=14)
bold12 = Font(bold=True, size=12)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, size=10, color='FFFFFF')
section_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
subtotal_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
total_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
num_fmt = '#,##0'


def styled_header(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin
        c.alignment = Alignment(horizontal='center')


def data_row(ws, row, values, bold=False):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = thin
        if bold:
            c.font = Font(bold=True)
        if isinstance(v, (int, float)):
            c.number_format = num_fmt
            c.alignment = Alignment(horizontal='right')


# ===== Sheet 1: 5월 정산 =====
ws1 = wb.active
ws1.title = '5월_정산'

ws1.merge_cells('A1:G1')
ws1['A1'] = '정비 비용정산서 (2026년 5월)'
ws1['A1'].font = bold14

info = {
    'A2': '정산일:', 'B2': '2026-05-03',
    'D2': '부서:', 'E2': '인프라팀',
    'A3': '설비코드:', 'B3': 'CH-001', 'C3': 'Chiller #1',
    'D3': '정비유형:', 'E3': '사후정비',
    'A4': '작성자:', 'B4': '홍길동',
    'D4': '승인자:', 'E4': '김부장',
    'F2': '공사번호:', 'G2': 'MO-2026-0503-001',
    'F3': '작업구역:', 'G3': '유틸동 B1',
}
for pos, val in info.items():
    ws1[pos] = val
    if str(val).endswith(':'):
        ws1[pos].font = Font(bold=True, size=10, color='333333')

# 자재비 섹션
r = 6
ws1.cell(row=r, column=1, value='[자재비]').font = bold12
for col in range(1, 8):
    ws1.cell(row=r, column=col).fill = section_fill
r += 1

styled_header(ws1, r, ['No.', '자재명', '규격/사양', '수량', '단가(원)', '금액(원)', '비고'])
r += 1

materials = [
    (1, '베어링 6205-2RS', 'NSK / 내경25mm', 4, 15000, 60000, '정기교체'),
    (2, 'V벨트 B-52', '반도 / 1321mm', 2, 8500, 17000, ''),
    (3, '메카니컬씰 MG1-35', 'Burgmann / 35mm', 1, 285000, 285000, '긴급 수입'),
    (4, '윤활유 Mobil SHC 630', 'ISO VG 220 / 20L', 1, 350000, 350000, '단가확인 필요'),
    (5, '가스켓 Non-Asbestos', 'Klinger / DN50', 6, 12000, 72000, ''),
    (6, '커플링 L-100', 'Lovejoy / 1-1/8"', 2, 45000, 90000, ''),
    (7, '베어링 6205-2RS', 'NSK / 내경25mm', 4, 15000, 60000, '추가 요청'),
    (8, '오링 P-30', 'NOK / Viton', 10, 3500, 35000, ''),
    (9, '임펠러 SUS316', '주문제작 / dia150', 1, 1200000, 1200000, '특별 발주'),
    (10, '볼트/너트 세트', 'STS304 M12x50', 20, 2500, 50000, ''),
]
for item in materials:
    data_row(ws1, r, list(item))
    r += 1

# 자재비 소계 (오류: 2,000,000 != 실제 2,219,000)
data_row(ws1, r, ['', '소계', '', '', '', 2000000, ''], bold=True)
for col in range(1, 8):
    ws1.cell(row=r, column=col).fill = subtotal_fill
r += 2

# 인건비 섹션
ws1.cell(row=r, column=1, value='[인건비]').font = bold12
for col in range(1, 8):
    ws1.cell(row=r, column=col).fill = section_fill
r += 1

styled_header(ws1, r, ['No.', '구분', '성명/업체', '투입시간(h)', '단가(원/h)', '금액(원)', '비고'])
r += 1

labors = [
    (1, '내부', '김철수', 8, 30000, 240000, '주간'),
    (2, '내부', '이영희', 10, 30000, 300000, '야간 포함'),
    (3, '외주', '(주)정비테크 박기사', 26, 45000, 1170000, '긴급 투입'),
    (4, '외주', '(주)정비테크 최기사', 16, 45000, 720000, ''),
    (5, '내부', '박인프라', 8, 30000, 240000, '지원'),
    (6, '외주', '(주)삼성ENG', 12, 65000, 780000, '전문업체'),
    (7, '내부', '정유틸', 6, 30000, 180000, ''),
]
for item in labors:
    data_row(ws1, r, list(item))
    r += 1

# 인건비 소계 (오류: 3,200,000 != 실제 3,630,000)
data_row(ws1, r, ['', '소계', '', '', '', 3200000, ''], bold=True)
for col in range(1, 8):
    ws1.cell(row=r, column=col).fill = subtotal_fill
r += 2

# 경비 섹션
ws1.cell(row=r, column=1, value='[경비]').font = bold12
for col in range(1, 8):
    ws1.cell(row=r, column=col).fill = section_fill
r += 1

styled_header(ws1, r, ['No.', '구분', '내용', '', '금액(원)', '', '비고'])
r += 1

expenses = [
    (1, '운반비', '크레인 임대 (25T)', '', 450000, '', '하루'),
    (2, '장비임대', '고소작업대 GENIE', '', 280000, '', '2일'),
    (3, '폐기물', '폐유/폐자재 처리', '', 150000, '', ''),
    (4, '기타', '안전장구 구매', '', 85000, '', ''),
]
for item in expenses:
    data_row(ws1, r, list(item))
    r += 1

data_row(ws1, r, ['', '소계', '', '', 965000, '', ''], bold=True)
for col in range(1, 8):
    ws1.cell(row=r, column=col).fill = subtotal_fill
r += 2

# 총합계 (오류)
ws1.cell(row=r, column=1, value='총 합계').font = bold12
for col in range(1, 8):
    ws1.cell(row=r, column=col).fill = total_fill
c = ws1.cell(row=r, column=5, value=5800000)
c.font = bold12
c.number_format = num_fmt
c.border = thin

widths = {'A': 6, 'B': 24, 'C': 22, 'D': 14, 'E': 14, 'F': 14, 'G': 18}
for col, w in widths.items():
    ws1.column_dimensions[col].width = w


# ===== Sheet 2: 4월 정산 =====
ws2 = wb.create_sheet('4월_정산')
ws2['A1'] = '정비 비용정산서 (2026년 4월)'
ws2['A1'].font = bold14
ws2['A2'] = '정산일:'
ws2['B2'] = '2026-04-15'
ws2['D2'] = '부서:'
ws2['E2'] = '인프라팀'
ws2['A3'] = '설비코드:'
ws2['B3'] = 'CH-001'
ws2['D3'] = '정비유형:'
ws2['E3'] = '예방정비'

styled_header(ws2, 5, ['No.', '자재명', '규격/사양', '수량', '단가(원)', '금액(원)', '비고'])
apr_mats = [
    (1, '윤활유 Mobil SHC 630', 'ISO VG 220 / 20L', 1, 112000, 112000, ''),
    (2, '필터 엘리먼트', 'Pall HC8314FKP16Z', 2, 85000, 170000, ''),
    (3, '가스켓 Non-Asbestos', 'Klinger / DN50', 4, 12000, 48000, ''),
]
r2 = 6
for item in apr_mats:
    data_row(ws2, r2, list(item))
    r2 += 1
data_row(ws2, r2, ['', '소계', '', '', '', 330000, ''], bold=True)

for col, w in widths.items():
    ws2.column_dimensions[col].width = w


# ===== Sheet 3: 단가기준표 =====
ws3 = wb.create_sheet('단가기준표')
ws3['A1'] = '자재 단가 기준표 (2026)'
ws3['A1'].font = bold14

styled_header(ws3, 3, ['자재명', '규격', '기준단가', '최저가', '최고가', '비고', ''])
prices = [
    ('베어링 6205-2RS', 'NSK', 15000, 12000, 18000, '', ''),
    ('V벨트 B-52', '반도', 8500, 7000, 10000, '', ''),
    ('메카니컬씰 MG1-35', 'Burgmann', 285000, 250000, 320000, '수입품', ''),
    ('윤활유 Mobil SHC 630', '20L', 112000, 95000, 130000, '대리점 가격', ''),
    ('가스켓 Non-Asbestos', 'Klinger', 12000, 10000, 15000, '', ''),
    ('커플링 L-100', 'Lovejoy', 45000, 38000, 52000, '', ''),
    ('오링 P-30', 'NOK', 3500, 2800, 4200, '', ''),
    ('임펠러 SUS316', '주문제작', 900000, 800000, 1100000, '사양별 상이', ''),
]
r3 = 4
for item in prices:
    data_row(ws3, r3, list(item))
    r3 += 1

for col, w in {'A': 24, 'B': 16, 'C': 14, 'D': 14, 'E': 14, 'F': 14, 'G': 4}.items():
    ws3.column_dimensions[col].width = w


path = 'C:/PP_CC_Error/sample_settlement.xlsx'
wb.save(path)
print(f'Created: {path}')
print('3 sheets: 5월_정산, 4월_정산, 단가기준표')
print()
print('Errors:')
print('1. 자재비 소계 2,000,000 != 실제 2,219,000')
print('2. 인건비 소계 3,200,000 != 실제 3,630,000')
print('3. 베어링 6205-2RS 중복 청구')
print('4. 박기사 26h (24h 초과)')
print('5. 윤활유 350,000원 (기준 112,000원)')
print('6. 총합계 5,800,000 != 소계합 6,165,000')
